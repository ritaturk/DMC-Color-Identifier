import openpyxl
from Utils.Color import rgba2hex


def font_color(hex_color):
    """
    Determines the font color based on the RGB value.
    
    Parameters:
    - hex_color: Hexadecimal color string.
    
    Returns:
    - Font color ('FFFFFF' for white or '000000' for black).
    """
    rgb_hex = hex_color[2:]  # Extract RGB portion from the hex color
    return 'FFFFFF' if int(rgb_hex, 16) < 0x7FFFFF else '000000'  # Return white if dark, black if light


def createExcel(index_matrix, real_matrix, dmc_matrix, sprite_colors, file_path):
    """
    Creates an Excel file with the color matrices and associated data.
    
    Parameters:
    - index_matrix: Matrix of color indices.
    - real_matrix: Matrix of real colors in hex.
    - dmc_matrix: Matrix of DMC colors in hex.
    - sprite_colors: DataFrame of sprite colors.
    - filename: Name of the Excel file to be created.
    """
    
    
    wb = openpyxl.Workbook()  # Create a new Excel workbook
    ws_real = wb.active  # Get the active worksheet
    ws_real.title = 'REAL'  # Set worksheet title
    
    for i in range(index_matrix.shape[0]):  # Iterate over rows of the index matrix
        for j in range(index_matrix.shape[1]):  # Iterate over columns
            if index_matrix[i, j] == 'x':  # Skip if index is 'x'
                continue
            # Set cell value and background color based on real matrix
            cell = ws_real.cell(row=i+1, column=j+1, value=index_matrix[i, j])
            cell.fill = openpyxl.styles.PatternFill(start_color=real_matrix[i, j], end_color=real_matrix[i, j], fill_type='solid')            
            cell.font = openpyxl.styles.Font(color=font_color(real_matrix[i, j]))  # Set font color

    # Add additional columns for Palette and Hexadecimal
    palette_start = index_matrix.shape[1]+2
    ws_real.cell(row=1, column=palette_start, value='Palette')
    ws_real.cell(row=1, column=palette_start+1, value='Hexadecimal')
    # Adjust column widths for better visibility
    ws_real.column_dimensions[openpyxl.utils.get_column_letter(palette_start)].width = 15
    ws_real.column_dimensions[openpyxl.utils.get_column_letter(palette_start+1)].width = 15

    # Fill in color palette data in additional columns
    for i in range(2, len(sprite_colors)+2):
        ws_real.cell(row=i, column=palette_start, value=sprite_colors.loc[i-2, 'INDEX'])
        ws_real.cell(row=i, column=palette_start+1, value=f"#{rgba2hex(sprite_colors.loc[i-2, 'REAL'])}")
        ws_real.cell(row=i, column=palette_start).fill = openpyxl.styles.PatternFill(start_color=rgba2hex(sprite_colors.loc[i-2, 'REAL']), end_color=rgba2hex(sprite_colors.loc[i-2, 'REAL']), fill_type='solid')
        ws_real.cell(row=i, column=palette_start).font = openpyxl.styles.Font(color=font_color(rgba2hex(sprite_colors.loc[i-2, 'REAL'])))
        ws_real.cell(row=i, column=palette_start+1).fill = openpyxl.styles.PatternFill(start_color=rgba2hex(sprite_colors.loc[i-2, 'REAL']), end_color=rgba2hex(sprite_colors.loc[i-2, 'REAL']), fill_type='solid')
        ws_real.cell(row=i, column=palette_start+1).font = openpyxl.styles.Font(color=font_color(rgba2hex(sprite_colors.loc[i-2, 'REAL'])))

    ws_dmc = wb.create_sheet(title='DMC')  # Create a new worksheet for DMC colors
    for i in range(index_matrix.shape[0]):  # Iterate over index matrix
        for j in range(index_matrix.shape[1]):
            if index_matrix[i, j] == 'x':  # Skip if index is 'x'
                continue
            # Set cell value and background color based on DMC matrix
            cell = ws_dmc.cell(row=i+1, column=j+1, value=index_matrix[i, j])
            cell.fill = openpyxl.styles.PatternFill(start_color=dmc_matrix[i, j], end_color=dmc_matrix[i, j], fill_type='solid')
            cell.font = openpyxl.styles.Font(color=font_color(dmc_matrix[i, j]))  # Set font color

    # Add additional columns for DMC details
    ws_dmc.cell(row=1, column=palette_start, value='Palette')
    ws_dmc.cell(row=1, column=palette_start+1, value='Floss')
    ws_dmc.cell(row=1, column=palette_start+2, value='Hexadecimal')
    # Adjust column widths for better visibility
    ws_dmc.column_dimensions[openpyxl.utils.get_column_letter(palette_start)].width = 15
    ws_dmc.column_dimensions[openpyxl.utils.get_column_letter(palette_start+1)].width = 15
    ws_dmc.column_dimensions[openpyxl.utils.get_column_letter(palette_start+2)].width = 15

    # Fill in DMC details in additional columns
    for i in range(2, len(sprite_colors)+2):
        ws_dmc.cell(row=i, column=palette_start, value=sprite_colors.loc[i-2, 'INDEX'])
        ws_dmc.cell(row=i, column=palette_start+1, value=sprite_colors.loc[i-2, 'FLOSS'])
        ws_dmc.cell(row=i, column=palette_start+2, value=f"#{rgba2hex(sprite_colors.loc[i-2, 'DMC'])}")        
        ws_dmc.cell(row=i, column=palette_start).fill = openpyxl.styles.PatternFill(start_color=rgba2hex(sprite_colors.loc[i-2, 'DMC']), end_color=rgba2hex(sprite_colors.loc[i-2, 'DMC']), fill_type='solid')
        ws_dmc.cell(row=i, column=palette_start).font = openpyxl.styles.Font(color=font_color(rgba2hex(sprite_colors.loc[i-2, 'DMC'])))
        ws_dmc.cell(row=i, column=palette_start+1).fill = openpyxl.styles.PatternFill(start_color=rgba2hex(sprite_colors.loc[i-2, 'DMC']), end_color=rgba2hex(sprite_colors.loc[i-2, 'DMC']), fill_type='solid')
        ws_dmc.cell(row=i, column=palette_start+1).font = openpyxl.styles.Font(color=font_color(rgba2hex(sprite_colors.loc[i-2, 'DMC'])))
        ws_dmc.cell(row=i, column=palette_start+2).fill = openpyxl.styles.PatternFill(start_color=rgba2hex(sprite_colors.loc[i-2, 'DMC']), end_color=rgba2hex(sprite_colors.loc[i-2, 'DMC']), fill_type='solid')
        ws_dmc.cell(row=i, column=palette_start+2).font = openpyxl.styles.Font(color=font_color(rgba2hex(sprite_colors.loc[i-2, 'DMC'])))

    # Adjust row heights and column widths for better layout
    for i in range(1, index_matrix.shape[1] + 1):
        ws_real.row_dimensions[i].height = 15  # Set row height for REAL sheet
        ws_real.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 3  # Set column width for REAL sheet
        ws_dmc.row_dimensions[i].height = 15  # Set row height for DMC sheet
        ws_dmc.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 3  # Set column width for DMC sheet

    # Add borders to all cells in the workbook
    for ws in wb:
        for row in ws.iter_rows():  # Iterate over each row in the worksheet
            for cell in row:  # Iterate over each cell in the row
                cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), 
                                                      right=openpyxl.styles.Side(style='thin'), 
                                                      top=openpyxl.styles.Side(style='thin'), 
                                                      bottom=openpyxl.styles.Side(style='thin'))  # Set cell borders

    # Center align text in all cells
    for ws in wb:
        for row in ws.iter_rows():  # Iterate over each row in the worksheet
            for cell in row:  # Iterate over each cell in the row
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')  # Center align cell text

    # Save the workbook to the specified results folder
    wb.save(file_path)
    print(f"\tExcel file saved to {file_path}")  # Print confirmation message